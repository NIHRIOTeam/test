# -*- coding: utf-8 -*-
"""
Created on Fri Jan 25 11:27:43 2019

@author: nsp133
"""

import datetime
import difflib
import urllib
import urllib.request
import tqdm
import time
import openpyxl

#from Grant.Wellcome import MongoConfig
wb = openpyxl.load_workbook('D:\\MC_Aging\\pubmed_result.xlsx')
#D:\\Grants\\Awards File for MRC website updated 10_05_2018 with ODA details.xlsx
#file = 'https://mrc.ukri.org/documents/xls-csv/grants-and-fellowships-awarded-xlsx/'
headList=[]
row_count=0
class MRC():
    
    def MRC(wb):
        
        
        #print('sheet names',wb.sheetnames[0])
        sheet = wb.get_sheet_by_name(wb.sheetnames[0])
        row_count = sheet.max_row  # count number of rows
        print(row_count)
        column_count = sheet.max_column  #count number of columns
        # k=0
        for i in range(1,2):  #get the first row as heading
            head = []
            headValue=''
            for ch in (map(chr, range(65,76))):  #iterate over the columns A : Z
                head.append(ch + str(1))
            for m in head:                 # iterate over the list and generate the full row
                headValue = str(sheet[m].value)
                headList.append(headValue.replace(';',''))
        print(headList)
        for i in tqdm.tqdm(range(2,row_count+1)):  #iterate over the number of rows not on the heading
            col=[]
            colList=[]
            lst = []
            Title=URL=Description=ShortDetails=Resources=Type=Identifiers=Db=EntrezUID=Properties=''
            for j in (map(chr, range(65,76))):  #iterate over the columns A : Z
                col.append(j + str(i))   #A[1]  create list of first row
            for k in col:                 # iterate over the list and generate the full row
                colValue = str(sheet[k].value)
                colList.append(colValue.replace(';',''))
            print(colList)
            f = zip(headList,colList)
            file = open('D:\\MC_Aging\\Results_split\\PMID_' + str(k) + ".txt", "w+")
            for key,value in f:
                if 'Title' in key: #
                    Abstract=  value
                    value1= ("Title" +" =: " + value)
                    print(value1)
                elif 'URL' in key: #
                    GrantID=  value
                    value1= ("URL" +" =: " + 'https://www.ncbi.nlm.nih.gov' + value)
                elif 'Description' in key: #
                    if value !='None':
                        Investigator=  value
                        value1= ("Description" +" =: " + value)
                elif 'ShortDetails' in key: #
                    if value !='None':
                        Scheme=  value
                        dic={'Access to Fellowships':'Fellowship','Access to Research Grants':'Research Grant','Access to Studentships':'Training Grant','BMC: DPFS Full':'Research Grant','Career Development Award Pre-F':'Fellowship','Career Development Award':'Fellowship','Career Establishment Grant':'Research Grant','Centre Component Grant':'Centre','Centre Development Grant PreF':'Centre','Centre Grant Pre FEC':'Research Grant','Centres':'Centre','Clinical Res Professorship Pre':'Fellowship','Clinical Res Train Fell Non-Fe':'Fellowship','Clinical Res Train Fell Pre-Fe':'Fellowship','Clinical Res Trn Fellowship':'Fellowship','Clinician Scientist Fell Pre-F':'Fellowship','Clinician Scientist Fellowship':'Fellowship','Co-operative Group Component G':'Research Grant','Co-operative Group Grant':'Research Grant','Collab Career Dev Fellow Pre F':'Fellowship','Collaboration grants':'Research Grant','Developmental Clinical Studies':'Research Grant','Developmental Pathway Funding':'Research Grant','Discipline Hopping Awards':'Research Grant','Doctoral Training Partnerships':'Training Grant','Equipment Project grant':'Research Grant','FEC':'Fellowship','Fut Lead  Ageing Res Fellows':'Fellowship','Institutional Discipline Bridg':'Institutional','Joint Funded Initiatives Full':'Research Grant','LINK GRANT Pre Fec':'Research Grant','LINK GRANT':'Research Grant','MRC Capacity Building':'Training Grant','MRC Doctoral Training Grant':'Training Grant','MRC Industrial CASE':'Training Grant','MRC/ESRC Interdisc Postdoc Fel':'Fellowship','NPRI Grant':'Research Grant','New Investigator Awards PreFec':'Research Grant','New Investigator Awards':'Research Grant','Newton Fund':'Innovation (industry)','Non Clin Res Professorship Pre':'Fellowship','Non-Clinical Res Readership Pr':'Fellowship','P&Cs':'Other','People Exchange Fellowship':'Fellowship','Programme Grant Pre FEC':'Research Grant','Programme Grant':'Research Grant','Programme grant (small)':'Research Grant','RCUK':'Research Grant','Regen Med Prog':'Research Grant','Research /Programme grant':'Research Grant','Research Fellowship Pre FEC':'Fellowship','Research Fellowship':'Fellowship','Research Grant':'Research Grant','Research Grants Pre Fec':'Research Grant','Research Grants':'Research Grant','Research/Programme grant':'Research Grant','Senior Clinical Fellowship Pre':'Fellowship','Senior Clinical Fellowship':'Fellowship','Senior Non-Clinical Fell Pre F':'Fellowship','Senior Non-Clinical Fellowship':'Fellowship','Skills Development Fellowships':'Fellowship','Sp Trn Fellowship Bioinformat':'Fellowship','Spec Train Fellowship(SPECTRN)':'Fellowship','Special Train Fell in Heal Pre':'Fellowship','Special Train Fellow in Comp B':'Fellowship','Strategic Appointments Scheme':'Strategic','Strategic Awards':'Strategic','Strategic Grant Pre Fec':'Strategic','Strategic Grant2':'Strategic','Strategic Project Grant (Trial':'Strategic','Strategic Project grant':'Strategic','Total Cost Fellowships':'Fellowship','Trial Grant Pre Fec':'Research Grant','Trial Grant':'Research Grant','Unit - ESS':'Unit','Unit - University Unit':'Unit','Unit':'Unit'}
                        #a =[value]
                        #scheme = ([dic.get(n, n) for n in a])
                        #print(scheme[0])
                        value1= ("ShortDetails" +" =: " + value)
               
                elif 'Resource' in key: #
                    if value !='None':
                        Start=  value
                        #date_datestart = datetime.datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
                        value1= ("Resource" +" =: " + value)
                elif 'Type' in key: #
                    if value !='None':
                        End=  value
                        #date_date = datetime.datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
                        value1= ("Type" +" =: " + value)
                elif 'Identifiers' in key: #
                   if value !='None':
                        #Amount=  float(value)
                        value1= ("Identifiers" +" =: " + value)
                elif 'Db' in key: #
                    if value !='None':
                        TopicArea=  value
                        value1= ("BD" +" =: " + value)
                elif 'EntrezUID' in key:   #
                    if value !='None':
                        Title=  value
                        value1= ("EntrezUID" +" =: " + value)
                elif 'Properties' in key:   #
                    if value !='None':
                        #Department=  value
                        value1= ("Properties" +" =: " + value)
                        print(value1)
                else:
                    key=  value
                    value1= (key +" =: " + value)
                file.write(value1.encode('cp850','replace').decode('cp850'))
                file.write('\n')
            #MongoConfig.db.GrantMRC.insert({"TopicArea": TopicArea,"FundingBody": "MRC","NormalizedGrantTypeScheme": scheme[0],"OriginalGrantTypeScheme": Scheme,"GrantID": GrantID,"StartDate": date_datestart, "EndDate": date_date,"Abstract": Abstract, "Title":Title, "Investigator":Investigator,"Department":Department,"Institution":Institution,"GrantAmount":Amount})
            file.close()
            time.sleep(0.01)
        print('done')
if __name__ == "__main__":
    #MRC.download(url)
    #wb = openpyxl.load_workbook('D:\\MC_Aging\\pubmed_result.xlsx')
    MRC.MRC(wb)

